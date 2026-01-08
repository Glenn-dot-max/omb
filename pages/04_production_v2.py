# pages/ 04_production.py
import streamlit as st
from datetime import date, timedelta, datetime
from collections import defaultdict
import sys
from pathlib import Path
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Ajouter le chemin parent pour importer database
sys.path.append(str(Path(__file__).parent.parent))

from database import (
  get_commandes,
  get_commande_details,
  get_categories,
  get_produit_by_id,
  get_produits_formule_avec_calcul,
  init_db_if_needed
)

# ========================================
# 🔐 PROTECTION PAR MOT DE PASSE
# ========================================
import sys
from pathlib import Path
sys.path.append(str(Path(__file__).parent.parent))

from auth import check_password

if not check_password():
    st.stop()

# Helper pour convertir les dates
def to_date(d):
    from datetime import date as date_type
    return d if isinstance(d, date_type) else date_type.fromisoformat(d)



# Configuration
st.set_page_config(page_title="Planning Production", page_icon="📊", layout="wide")

# Initialiser la base 
init_db_if_needed()

# Titre
st.title("📊 Planning de production")

# ======= FILTRES DE PÉRIODES ==========
st.subheader("Période de production")

col1, col2, col3 = st.columns([2, 2, 1])

with col1:
  date_debut = st.date_input(
    "Date de début",
    value=date.today(),
    key="date_debut_prod"
  )

with col2:
  # Sélection de vue
  vue = st.selectbox(
    "Vue",
    ["Jour", "3 jours", "Semaine", "2 semaines"],
    key="vue_prod"
  )

with col3:
  st.write("")
  st.write("")
  # Bouton pour générer le planning
  generer = st.button("🔄 Générer le planning", type="primary", use_container_width=True)
  st.session_state.generer_planning = True

# Calculer la date de fin selon la vue
if vue == "Jour":
  date_fin = date_debut
  nb_jours = 1
elif vue == "3 jours":
  date_fin = date_debut + timedelta(days=2)
  nb_jours = 3
elif vue == "Semaine":
  date_fin = date_debut + timedelta(days=6)
  nb_jours = 7
else:  # 2 semaines
  date_fin = date_debut + timedelta(days=13)
  nb_jours = 14

st.info(f"📅 Période: du {date_debut.strftime('%d/%m/%Y')} au {date_fin.strftime('%d/%m/%Y')}")

# ================= SELECTEUR DE PÉRIODE POUR TOTAUX =====================
st.subheader("📊 Configuration des totaux")

col_p1, col_p2, col_p3 = st.columns(3)

with col_p1:
  mode_total = st.selectbox(
    "Type de période",
    ["Jour par jour", "Semaine", "Jours personnalisés"],
    key="mode_total"
  )

with col_p2:
  if mode_total == "Jours personnalisés":
    jours_semaine = ["Lundi", "Mardi", "Mercredi", "Jeudi", "Vendredi", "Samedi", "Dimanche"]
    jours_selectionnes = st.multiselect(
      "Choisir les jours",
      jours_semaine,
      default=["Jeudi", "Vendredi"],
      key="jours_custom"
    )
  
with col_p3:
  afficher_total = st.checkbox("Afficher colonne Total", value=True, key="show_total")

st.divider()

if st.session_state.get('generer_planning', False):
  st.divider()

  # =============== RÉCUPÉRATION ET TRAITEMENT DES DONNÉES ===============

  # Récupérer toutes les commandes
  toutes_commandes = get_commandes()

  # Filtrer les commandes dans la période
  commandes_periode = []
  for cmd in toutes_commandes:
    cmd_date = to_date(cmd['date'])
    if date_debut <= cmd_date <= date_fin:
      commandes_periode.append({
        **cmd,
        'date_obj': cmd_date
      })

  # Trier par date et heure
  commandes_periode = sorted(commandes_periode, key=lambda x: (x['date_obj'], x['heure']))

  if not commandes_periode:
    st.warning("Aucune commande dans cette période")
    st.stop()

  st.success(f"✅ {len(commandes_periode)} commande(s) trouvée(s) dans la période")

  # =============== ORGANISER LES DONNÉES ============

  # Structure : {jour: {commande_id: {nom_client, produits}}}
  planning = defaultdict(lambda: defaultdict(dict))

  # Structure : {categorie: {produit_nom: {jour: {commande_id: quantité}}}}
  produits_par_categorie = defaultdict(lambda: defaultdict(lambda: defaultdict(lambda: defaultdict(dict))))

  # Liste de tous les produits avec leur catégorie
  tous_produits = set()

  for cmd in commandes_periode:
    jour = cmd['date_obj'].strftime('%Y-%m-%d')
    jour_affichage = cmd['date_obj'].strftime('%d/%m')

    # Récupérer le détail de la commande
    details = get_commande_details(cmd['id'])

    # Stocker les infos de la commande
    planning[jour][cmd['id']] = {
      'client': cmd['client'],
      'couverts': cmd['couverts'],
      'heure': cmd['heure'],
      'jour_affichage': jour_affichage
    }

    # ========== TRAITER LES PRODUITS DES FORMULES ==========
    if details['formules']:
        for formule_id, formule_nom, qte_recommandee, qte_finale in details['formules']:
            # Récupérer les produits de cette formule avec calcul
            produits_formule = get_produits_formule_avec_calcul(formule_id, cmd['couverts'])
            
            for prod in produits_formule:
                # Les vraies clés sont : id, nom, qte_recommandee, unite
                produit_id = prod['id']
                prod_nom = prod['nom']
                qte = prod['qte_recommandee']
                unite_nom = prod['unite']
                
                # Récupérer la catégorie du produit
                produit_info = get_produit_by_id(produit_id)
                categorie = produit_info['categorie'] if produit_info and produit_info.get('categorie') else "Sans catégorie"
                
                # Ajouter ou cumuler la quantité
                if cmd['id'] in produits_par_categorie[categorie][prod_nom][jour]:
                    # Cumuler si le même produit existe déjà
                    produits_par_categorie[categorie][prod_nom][jour][cmd['id']]['quantite'] += qte
                else:
                    # Initialiser sinon
                    produits_par_categorie[categorie][prod_nom][jour][cmd['id']] = {
                        'quantite': qte,
                        'unite': unite_nom or '',
                        'source': 'formule'
                    }
                
                tous_produits.add((categorie, prod_nom))

    # ========== TRAITER LES PRODUITS SUPPLÉMENTAIRES ==========
    if details['produits']:
      for produit_id, prod_nom, qte, unite_nom, unite_id in details['produits']:
        # Récupérer la catégorie du produit 
        produit_info = get_produit_by_id(produit_id)
        categorie = produit_info['categorie'] if produit_info and produit_info.get('categorie') else "Sans catégorie"

        # Ajouter ou cumuler la quantité
        if cmd['id'] in produits_par_categorie[categorie][prod_nom][jour]:
          # Cumuler si le même produit existe déjà
          produits_par_categorie[categorie][prod_nom][jour][cmd['id']]['quantite'] += qte
        else:
          produits_par_categorie[categorie][prod_nom][jour][cmd['id']] = {
            'quantite': qte,
            'unite': unite_nom or '',
            'source': 'suppl'
          }
        
        tous_produits.add((categorie, prod_nom))


  # ================= AFFICHAGE DU TABLEAU ==============

  st.subheader("📋 Plan de production")

  # Créer la liste des jours dans la période 
  jours_liste = []
  date_courante = date_debut
  jours_fr = ['Lundi', 'Mardi', 'Mercredi', 'jeudi', 'Vendredi', 'Samedi', 'Dimanche']

  for i in range(nb_jours):
    jour_str = date_courante.strftime('%Y-%m-%d')
    jour_num = date_courante.strftime('%d/%m')
    jour_semaine = jours_fr[date_courante.weekday()]
    jours_liste.append({
      'date': jour_str,
      'affichage': f"{jour_num}<br/>{jour_semaine}",
      'nom_jour': jour_semaine
    })
    date_courante += timedelta(days=1)

  # Afficher par catégorie
  categories = sorted(produits_par_categorie.keys())

  if not categories:
    st.warning("⚠️ Aucun produit trouvé dans les commandes de cette période.")
    st.stop()

  # Pour chaque catégorie
  for categorie in categories:
    with st.expander(f"📦 {categorie}", expanded=True):
      
      # Créer le tableau HTML
      html_table = "<table style='width:100%; border-collapse: collapse; font-size: 11px;'>"

      # En-tête : Jours avec colonnes TOTAL
      html_table += "<thead><tr style='background-color: #4472C4; color: white; font-weight: bold;'>"
      html_table += "<th style='border: 1px solid #ddd; padding: 8px; text-align: left; min-width: 180 px;'>Produit</th>"

      for jour_info in jours_liste:
        jour = jour_info['date']
        nom_jour = jour_info['nom_jour']

        # Formater la date pour l'affichage 
        date_obj = datetime.strptime(jour, '%Y-%m-%d')        
        jour_display = f"{date_obj.strftime('%d/%m')}<br/>{nom_jour}"

        # Compter le nombre de commandes ce jour-là
        nb_commandes = len(planning.get(jour, {}))
        if nb_commandes > 0:
          # Créer une colonne par commande + 1 colonne TOTAL
          html_table += f"<th colspan='{nb_commandes + 1}' style='border: 1px solid #ddd; padding: 8px; text-align: center; background-color: #5B9BD5;'>{jour_display}</th>"
        else:
          html_table += f"<th colspan='2' style='border: 1px solid #ddd; padding: 8px; text-align: center; background-color: #D9E1F2;'>{jour_display}</th>"
      # Colonne TOTAL GÉNÉRAL
      html_table += "<th style='border: 1px solid #ddd; padding: 8px; text-align: center; background-color: #ED7D31; color: white;'>TOTAL<br/>GÉNÉRAL</th>"

      # Sous-en-tête : Noms des clients
      html_table += "<tr style='background-color: #D9E1F2;'>"
      html_table += "<th style='border: 1px solid #ddd; padding: 4px; font-weight: normal; font-style: italic;'>Client</th>"

      for jour_info in jours_liste:
        jour = jour_info['date']
        nom_jour = jour_info['nom_jour']
        commandes_jour = planning.get(jour, {})

        if commandes_jour:
          for cmd_id, cmd_info in commandes_jour.items():
            client_short = cmd_info['client'][:20] # Limiter à 20 caractères
            html_table += f"<th style='border: 1px solid #ddd; padding: 4px; font-size: 9 px; text-align: center;'>{client_short}<br/><span style='color: #666;'>{cmd_info['heure']}</span></th>"
          # Colonne TOTAL du jour
          html_table += f"<th style='border: 1px solid #ddd; padding: 4px; font-weight: bold; text-align: center; background-color: #F4B084;'>TOTAL<br/>{nom_jour}</th>"
        else:
          html_table += "<th style='border: 1px solid #ddd; padding: 4px;'>-</th>"
          html_table += f"<th style='border: 1px solid #ddd; padding: 4px; font-weight: bold; background-color: #F4B084;'>TOTAL<br/>{nom_jour}</th>"

      # Sous-en-tête pour TOTAL GÉNÉRAL
      html_table += "<th style='border: 1px solid #ddd; padding: 4px; background-color: #F4B084;'></th>"
      html_table += "</tr></thead>"

      # Corps du tableau: Produits
      html_table += "<tbody>"

      produits_categorie = sorted(produits_par_categorie[categorie].keys())

      for prod_nom in produits_categorie:
              html_table += "<tr>"
              html_table += f"<td style='border: 1px solid #ddd; padding: 8px; font-weight: bold; background-color: #FFF2CC;'>{prod_nom}</td>"

              # Total général pour ce produit
              total_general = 0
              unite_produit = ""
              
              for jour_info in jours_liste:
                jour = jour_info['date']
                commandes_jour = planning.get(jour, {})

                # Total du jour pour ce produit
                total_jour = 0

                if commandes_jour: 
                  for cmd_id in commandes_jour.keys():
                    # Vérifier si ce produit est dans cette commande 
                    if cmd_id in produits_par_categorie[categorie][prod_nom].get(jour, {}):
                      info = produits_par_categorie[categorie][prod_nom][jour][cmd_id]
                      qte = info['quantite']
                      unite = info['unite']
                      unite_produit = unite  # Garder l'unité pour le total
                      total_jour += qte
                      total_general += qte
                      
                      # Couleur selon la source
                      bg_color = "#E8F5E9" if info['source'] == 'formule' else "#FFF3E0"
                      
                      html_table += f"<td style='border: 1px solid #ddd; padding: 6px; text-align: center; background-color: {bg_color};'>{int(qte) if qte == int(qte) else qte:.1f}</td>"
                    else:
                      html_table += "<td style='border: 1px solid #ddd; padding: 6px; text-align: center; background-color: #f9f9f9;'>-</td>"
                  
                  # Afficher le TOTAL du jour
                  if total_jour > 0:
                    html_table += f"<td style='border: 1px solid #ddd; padding: 6px; text-align: center; font-weight: bold; background-color: #FFE699;'>{int(total_jour) if total_jour == int(total_jour) else total_jour:.1f}</td>"
                  else:
                    html_table += "<td style='border: 1px solid #ddd; padding: 6px; text-align: center; background-color: #f9f9f9;'>-</td>"
                else:
                  html_table += "<td style='border: 1px solid #ddd; padding: 6px; text-align: center; background-color: #f9f9f9;'>-</td>"
                  html_table += "<td style='border: 1px solid #ddd; padding: 6px; text-align: center; background-color: #f9f9f9;'>-</td>"
              
              # Afficher le TOTAL GÉNÉRAL
              if total_general > 0:
                html_table += f"<td style='border: 1px solid #ddd; padding: 6px; text-align: center; font-weight: bold; background-color: #F8CBAD; font-size: 12px;'>{int(total_general) if total_general == int(total_general) else total_general:.1f} {unite_produit}</td>"
              else:
                html_table += "<td style='border: 1px solid #ddd; padding: 6px; text-align: center; background-color: #f9f9f9;'>-</td>"
            
              html_table += "</tr>"
          
      html_table += "</tbody></table>"

      # Afficher le tableau
      st.markdown(html_table, unsafe_allow_html=True)
      
      # Légende
      col1, col2 = st.columns(2)
      with col1:
        st.markdown("🟢 **Vert clair** : Produits des formules")
      with col2:
        st.markdown("🟠 **Orange clair** : Produits supplémentaires")

      st.write("")

  st.divider()

  # ============== STATISTIQUES ==============
  st.subheader("📈 Statistiques")

  col1, col2, col3, col4 = st.columns(4)

  with col1:
    st.metric("📦 Commandes", len(commandes_periode))

  with col2:
    total_couverts = sum(cmd['couverts'] for cmd in commandes_periode)
    st.metric("👥 Couverts", total_couverts)

  with col3:
    nb_produits_uniques = len(tous_produits)
    st.metric("🥖 Produits différents", nb_produits_uniques)
  
  with col4:
    nb_categories = len(categories)
    st.metric("📁 Catégories", nb_categories)

  # Détail par catégorie
  st.divider()
  st.subheader("📊 Détail par catégorie")
  
  cols = st.columns(min(len(categories), 4))
  for idx, categorie in enumerate(categories):
    with cols[idx % 4]:
      nb_produits_cat = len(produits_par_categorie[categorie])
      st.metric(f"📦 {categorie}", f"{nb_produits_cat} produit(s)")
  
  # ============== EXPORT EXCEL ==============
  st.divider()
  st.subheader("📥 Exporter")
  
  col1, col2, col3 = st.columns([2, 1, 2])
  
  with col2:
    if st.button("📊 Générer le fichier en Excel", type="primary", use_container_width=True):
      from openpyxl.utils import get_column_letter
      
      # Créer le workbook
      wb = Workbook()
      ws = wb.active
      ws.title = "Planning Production"
      
      # Styles
      header_fill = PatternFill(start_color="8B4513", end_color="8B4513", fill_type="solid")
      header_font = Font(color="FFFFFF", bold=True, size=12)
      subheader_fill = PatternFill(start_color="FFE4B5", end_color="FFE4B5", fill_type="solid")
      subheader_font = Font(bold=True, size=10)
      product_fill = PatternFill(start_color="FFF8DC", end_color="FFF8DC", fill_type="solid")
      product_font = Font(bold=True, size=11)
      formule_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
      suppl_fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
      border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
      )
      
      current_row = 1
      
      # Titre principal (inclure colonnes TOTAL)
      max_cols_needed = 1 + sum([(len(planning.get(j['date'], {})) + 1) or 2 for j in jours_liste]) + 1  # +1 pour chaque TOTAL jour + 1 pour TOTAL GÉNÉRAL
      ws.merge_cells(f'A{current_row}:{get_column_letter(max_cols_needed)}{current_row}')
      cell = ws[f'A{current_row}']
      cell.value = f"Planning de Production - {date_debut.strftime('%d/%m/%Y')} au {date_fin.strftime('%d/%m/%Y')}"
      cell.font = Font(bold=True, size=14)
      cell.alignment = Alignment(horizontal='center')
      current_row += 2
      
      # ========== EN-TÊTE JOURS (UNE SEULE FOIS) ==========
      col_idx = 1
      cell = ws.cell(current_row, col_idx)
      cell.value = "Produit"
      cell.fill = header_fill
      cell.font = header_font
      cell.border = border
      cell.alignment = Alignment(horizontal='center', vertical='center')
      col_idx += 1
      
      # Colonnes des jours
      for jour_info in jours_liste:
        jour = jour_info['date']
        nom_jour = jour_info['nom_jour']
        commandes_jour = planning.get(jour, {})
        nb_commandes = len(commandes_jour) if commandes_jour else 0
        
        # Nombre de colonnes = clients + 1 TOTAL
        nb_cols_total = (nb_commandes + 1) if nb_commandes > 0 else 2
        
        # Fusionner les cellules pour le jour (inclure colonne TOTAL)
        if nb_cols_total > 1:
          ws.merge_cells(start_row=current_row, start_column=col_idx, 
                        end_row=current_row, end_column=col_idx + nb_cols_total - 1)
        
        cell = ws.cell(current_row, col_idx)
        jour_num = jour[8:10] + '/' + jour[5:7]
        cell.value = f"{jour_num}\n{nom_jour}"
        cell.fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True, size=11)
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        col_idx += nb_cols_total
      
      # Colonne TOTAL GÉNÉRAL
      cell = ws.cell(current_row, col_idx)
      cell.value = "TOTAL\nGÉNÉRAL"
      cell.fill = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
      cell.font = Font(color="FFFFFF", bold=True, size=11)
      cell.border = border
      cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
      
      current_row += 1
      
      # ========== SOUS-EN-TÊTE CLIENTS (UNE SEULE FOIS) ==========
      col_idx = 1
      cell = ws.cell(current_row, col_idx)
      cell.value = "Client"
      cell.fill = subheader_fill
      cell.font = Font(bold=True, italic=True, size=10)
      cell.border = border
      cell.alignment = Alignment(horizontal='center', vertical='center')
      col_idx += 1
      
      for jour_info in jours_liste:
        jour = jour_info['date']
        nom_jour = jour_info['nom_jour']
        commandes_jour = planning.get(jour, {})
        
        if commandes_jour:
          for cmd_id, cmd_info in commandes_jour.items():
            cell = ws.cell(current_row, col_idx)
            cell.value = f"{cmd_info['client'][:20]}\n{cmd_info['heure']}"
            cell.fill = subheader_fill
            cell.font = Font(size=9)
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            col_idx += 1
          
          # Colonne TOTAL du jour
          cell = ws.cell(current_row, col_idx)
          cell.value = f"TOTAL\n{nom_jour}"
          cell.fill = PatternFill(start_color="F4B084", end_color="F4B084", fill_type="solid")
          cell.font = Font(bold=True, size=9)
          cell.border = border
          cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
          col_idx += 1
        else:
          cell = ws.cell(current_row, col_idx)
          cell.value = "-"
          cell.fill = subheader_fill
          cell.border = border
          cell.alignment = Alignment(horizontal='center', vertical='center')
          col_idx += 1
          
          # Colonne TOTAL du jour (même si pas de commande)
          cell = ws.cell(current_row, col_idx)
          cell.value = f"TOTAL\n{nom_jour}"
          cell.fill = PatternFill(start_color="F4B084", end_color="F4B084", fill_type="solid")
          cell.font = Font(bold=True, size=9)
          cell.border = border
          cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
          col_idx += 1
      
      # Sous-en-tête TOTAL GÉNÉRAL
      cell = ws.cell(current_row, col_idx)
      cell.value = ""
      cell.fill = PatternFill(start_color="F4B084", end_color="F4B084", fill_type="solid")
      cell.border = border
      
      current_row += 1
      
      # ========== POUR CHAQUE CATÉGORIE (SANS EN-TÊTES) ==========
      for categorie in categories:
        # Titre de catégorie
        ws.merge_cells(f'A{current_row}:{get_column_letter(max_cols_needed)}{current_row}')
        cell = ws[f'A{current_row}']
        cell.value = f"📦 {categorie}"
        cell.font = Font(bold=True, size=13, color="8B4513")
        cell.fill = PatternFill(start_color="F5F5DC", end_color="F5F5DC", fill_type="solid")
        current_row += 1
        
        # Produits de la catégorie
        produits_categorie = sorted(produits_par_categorie[categorie].keys())
        
        for prod_nom in produits_categorie:
          col_idx = 1
          
          # Nom du produit
          cell = ws.cell(current_row, col_idx)
          cell.value = prod_nom
          cell.fill = product_fill
          cell.font = product_font
          cell.border = border
          col_idx += 1
          
          # Total général pour ce produit
          total_general = 0
          unite_produit = ""
          
          # Quantités par jour/commande
          for jour_info in jours_liste:
            jour = jour_info['date']
            commandes_jour = planning.get(jour, {})
            total_jour = 0
            
            if commandes_jour:
              for cmd_id in commandes_jour.keys():
                cell = ws.cell(current_row, col_idx)
                
                if cmd_id in produits_par_categorie[categorie][prod_nom].get(jour, {}):
                  info = produits_par_categorie[categorie][prod_nom][jour][cmd_id]
                  qte = info['quantite']
                  unite = info['unite']
                  unite_produit = unite
                  total_jour += qte
                  total_general += qte
                  
                  cell.value = int(qte) if qte == int(qte) else round(qte, 1)
                  cell.fill = formule_fill if info['source'] == 'formule' else suppl_fill
                  cell.font = Font(bold=True)
                else:
                  cell.value = "-"
                
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center')
                col_idx += 1
              
              # Colonne TOTAL du jour
              cell = ws.cell(current_row, col_idx)
              if total_jour > 0:
                cell.value = int(total_jour) if total_jour == int(total_jour) else round(total_jour, 1)
                cell.fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
                cell.font = Font(bold=True)
              else:
                cell.value = "-"
              cell.border = border
              cell.alignment = Alignment(horizontal='center', vertical='center')
              col_idx += 1
            else:
              cell = ws.cell(current_row, col_idx)
              cell.value = "-"
              cell.border = border
              cell.alignment = Alignment(horizontal='center', vertical='center')
              col_idx += 1
              
              # Colonne TOTAL du jour (vide si pas de commande)
              cell = ws.cell(current_row, col_idx)
              cell.value = "-"
              cell.border = border
              cell.alignment = Alignment(horizontal='center', vertical='center')
              col_idx += 1
          
          # Colonne TOTAL GÉNÉRAL
          cell = ws.cell(current_row, col_idx)
          if total_general > 0:
            val = int(total_general) if total_general == int(total_general) else round(total_general, 1)
            cell.value = f"{val} {unite_produit}"
            cell.fill = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")
            cell.font = Font(bold=True, color="000000")
          else:
            cell.value = "-"
          cell.border = border
          cell.alignment = Alignment(horizontal='center', vertical='center')
          
          current_row += 1
        
        current_row += 2  # Espace entre catégories
      
      # Ajuster la largeur des colonnes
      ws.column_dimensions['A'].width = 25
      for col in range(2, col_idx):
        col_letter = get_column_letter(col)
        ws.column_dimensions[col_letter].width = 12
      
      # Sauvegarder dans un buffer
      buffer = BytesIO()
      wb.save(buffer)
      buffer.seek(0)
      
      # Bouton de téléchargement
      filename = f"planning_production_{date_debut.strftime('%Y%m%d')}_{date_fin.strftime('%Y%m%d')}.xlsx"
      st.download_button(
        label="⬇️ Télécharger le fichier Excel",
        data=buffer,
        file_name=filename,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary",
        use_container_width=True
      )
      st.success("✅ Fichier Excel généré avec succès !")
  
  with col3:
    st.info("💡 Le fichier Excel reprend exactement le planning avec les couleurs et la mise en forme.")